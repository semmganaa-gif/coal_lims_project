# app/routes/main/hourly_report.py
# -*- coding: utf-8 -*-
"""
Цагийн тайлан илгээх (Hourly Report) — Excel template бөглөж имэйл илгээх
"""

import os
from io import BytesIO
from datetime import timedelta

from flask import flash, redirect, url_for, current_app
from flask_login import login_required, current_user
from flask_babel import lazy_gettext as _l
from flask_mail import Message
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from sqlalchemy import select

from app import db, mail
from app.constants import UserRole
from app.utils.decorators import role_required
from app.models import Sample, SystemSetting
from app.repositories import SystemSettingRepository
from app.utils.datetime import now_local
from app.utils.database import safe_commit
from app.services.analytics_service import (
    detect_anomalies, analyze_trends, calculate_quality_score,
    generate_insights, ANALYSIS_SPECS,
)

from . import main_bp


def get_report_email_recipients():
    """
    Тайлан илгээх имэйл хаягуудыг SystemSetting-ээс авах

    Returns:
        dict: {'to': [...], 'cc': [...]}
    """
    result = {'to': [], 'cc': []}

    to_setting = SystemSettingRepository.get('email', 'report_recipients_to')
    if to_setting and getattr(to_setting, 'is_active', True) and to_setting.value:
        result['to'] = [e.strip() for e in to_setting.value.split(',') if e.strip()]

    cc_setting = SystemSettingRepository.get('email', 'report_recipients_cc')
    if cc_setting and getattr(cc_setting, 'is_active', True) and cc_setting.value:
        result['cc'] = [e.strip() for e in cc_setting.value.split(',') if e.strip()]

    return result


@main_bp.route("/send-hourly-report")
@login_required
@role_required(UserRole.SENIOR.value, UserRole.ADMIN.value)
def send_hourly_report():
    """Цагийн тайлан илгээх - зөвхөн senior, admin"""
    try:
        current_app.logger.debug("HOURLY REPORT STARTED (FIXED POSITIONING)")

        # =========================================================================
        # 1. ТОХИРГОО
        # =========================================================================

        # --- 2 HOURLY MAPPING (Суурь мөрүүд) ---
        PF_MAPPING = { 'PF211': 20, 'PF221': 33, 'PF231': 46 }
        HCC_KEYWORDS = ['UHG MV', 'UHG HV', 'BN HV', 'BN SSCC']
        MIDD_KEYWORDS = ['BN MASHCC', 'BN MIDD', 'UHG MASHCC', 'UHG MIDD', 'MASHCC_2']

        SUFFIX_OFFSET_MAP = {
            'D1': 0, 'D2': 1, 'D3': 2, 'D4': 3, 'D5': 4, 'D6': 5,
            'N1': 6, 'N2': 7, 'N3': 8, 'N4': 9, 'N5': 10, 'N6': 11
        }

        # 2H Баганын дугаарууд
        COL_2H_NAME = 1   # A: Sample ID
        COL_2H_MT   = 3   # C
        COL_2H_MAD  = 4   # D
        COL_2H_AAD  = 5   # E
        COL_2H_AD   = 6   # F
        COL_2H_VAD  = 7   # G
        COL_2H_VDAF = 8   # H
        COL_2H_GI   = 15  # O

        # --- 4 HOURLY CONFIG ---
        ROW_BUCKETS_4H = {
            (8, 12):  87,  # 10:00
            (12, 16): 91,  # 14:00
            (16, 20): 95,  # 18:00
            (20, 24): 99,  # 22:00
            (0, 4):   103, # 02:00
            (4, 8):   107  # 06:00
        }

        CF_OFFSET_MAP = {
            # MOD I
            "CF501":       { "cols": {"Aad": 3, "FM": 5}, "offset": 0 },
            "CF502":       { "cols": {"Aad": 3, "FM": 5}, "offset": 1 },
            "CF601":       { "cols": {"Aad": 3, "FM": 5}, "offset": 2 },
            "CF602":       { "cols": {"Aad": 3, "FM": 5}, "offset": 2 },
            # MOD II
            "CF521":       { "cols": {"Aad": 11, "FM": 13}, "offset": 0 },
            "CF522":       { "cols": {"Aad": 11, "FM": 13}, "offset": 1 },
            "CF621":       { "cols": {"Aad": 11, "FM": 13}, "offset": 2 },
            "CF622":       { "cols": {"Aad": 11, "FM": 13}, "offset": 2 },
            # MOD III
            "CF541":       { "cols": {"Aad": 19, "FM": 21}, "offset": 0 },
            "CF542":       { "cols": {"Aad": 19, "FM": 21}, "offset": 1 },
            "CF641":       { "cols": {"Aad": 19, "FM": 21}, "offset": 2 },
            "CF642":       { "cols": {"Aad": 19, "FM": 21}, "offset": 2 },
        }

        # =========================================================================
        # 2. ЦАГ ХУГАЦАА
        # =========================================================================
        current_time = now_local()
        calc_time = current_time - timedelta(hours=2)
        report_hour = (calc_time.hour // 2) * 2
        report_dt = calc_time.replace(hour=report_hour, minute=0, second=0, microsecond=0)

        report_time_str = report_dt.strftime('%H:%M')
        report_date_str = report_dt.strftime('%Y.%m.%d')

        if report_dt.hour < 8:
            start_time = (report_dt - timedelta(days=1)).replace(hour=8, minute=0, second=0, microsecond=0)
        else:
            start_time = report_dt.replace(hour=8, minute=0, second=0, microsecond=0)

        end_time = current_time

        # =========================================================================
        # 3. EXCEL НЭЭХ
        # =========================================================================
        template_path = os.path.join(current_app.root_path, 'static', 'hourly_template.xlsx')
        if not os.path.exists(template_path):
            if os.path.exists(template_path + ".xlsx"):
                template_path += ".xlsx"
            else:
                flash(_l("Загвар файл олдсонгүй!"), "danger")
                return redirect(url_for('main.index'))

        with open(template_path, "rb") as f:
            output = BytesIO(f.read())

        wb = load_workbook(output)
        ws = wb.active

        center_align = Alignment(horizontal='center', vertical='center')
        font_reg = Font(name='Times New Roman', size=12)
        font_bold = Font(name='Times New Roman', size=12, bold=False)

        # -------------------------------------------------------------
        # HEADER & COUNTER
        # -------------------------------------------------------------
        setting_key = f"report_counter_{start_time.year}"
        last_update_key = f"last_update_{start_time.year}"

        setting_count = SystemSettingRepository.get('report_config', setting_key)
        setting_last = SystemSettingRepository.get('report_config', last_update_key)

        current_count = int(setting_count.value) if setting_count else 0
        last_updated_date = setting_last.value if setting_last else ""
        today_str = start_time.strftime('%Y-%m-%d')

        if report_time_str == "07:00" and last_updated_date != today_str:
            current_count += 1
            if not setting_count:
                db.session.add(SystemSetting(category='report_config', key=setting_key, value=str(current_count)))
            else:
                setting_count.value = str(current_count)
            if not setting_last:
                db.session.add(SystemSetting(category='report_config', key=last_update_key, value=today_str))
            else:
                setting_last.value = today_str
            safe_commit(error_msg="Тайлангийн тоолуур хадгалахад алдаа гарлаа")

        display_count = current_count if current_count > 0 else 1

        ws['E10'] = f"{start_time.year}_{display_count:03d}"
        ws['E10'].font = font_reg
        ws['E10'].alignment = Alignment(horizontal='left', vertical='center')

        ws['E11'] = start_time.strftime('%d-%b-%Y')
        ws['E11'].font = font_reg
        ws['E11'].alignment = Alignment(horizontal='left', vertical='center')

        ws['O3'] = f"№ {start_time.strftime('%y-%m-%d')}"
        ws['O3'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['O3'].alignment = Alignment(horizontal='right', vertical='center')

        ws['C4'] = start_time.strftime('%Y.%m.%d')
        ws['C4'].font = font_reg

        # =========================================================================
        # 4. ХЭСЭГ 1: 2 HOURLY ДЭЭЖҮҮД (FIXED POSITIONING)
        # =========================================================================
        samples_2h = list(db.session.execute(
            select(Sample).where(
                Sample.received_date >= start_time,
                Sample.received_date <= end_time,
                Sample.client_name == 'CHPP',
                Sample.sample_type.in_(['2 hourly', '2 Hourly']),
            )
        ).scalars().all())

        row_is_partial = {20: True, 33: True, 46: True, 59: False, 72: False}

        for s in samples_2h:
            code_upper = (s.sample_code or "").strip().upper()
            calc = s.get_calculations()

            # 1. СУУРЬ МӨРӨӨ ОЛОХ
            start_row = None
            for pf_key, r_num in PF_MAPPING.items():
                if pf_key in code_upper:
                    start_row = r_num
                    break

            if not start_row:
                for k in HCC_KEYWORDS:
                    if k in code_upper:
                        start_row = 59
                        break
                if not start_row and "UHG MASHCC" in code_upper and "MASHCC_2" not in code_upper:
                    start_row = 59

            if not start_row:
                for k in MIDD_KEYWORDS:
                    if k in code_upper:
                        start_row = 72
                        break

            if not start_row:
                start_row = 59

            # 2. ШИЛЖИЛТЭЭ ОЛОХ
            row_offset = 0
            for suffix, offset in SUFFIX_OFFSET_MAP.items():
                if f"_{suffix}" in code_upper or code_upper.endswith(suffix):
                    row_offset = offset
                    break

            final_row = start_row + row_offset

            # 3. БИЧИХ
            def w_cell(c, v, f=font_reg):
                cell = ws.cell(row=final_row, column=c, value=v)
                cell.alignment = center_align
                cell.font = f

            w_cell(COL_2H_NAME, s.sample_code, font_bold)

            w_cell(COL_2H_MT, calc.mt if calc.mt is not None else "-")
            w_cell(COL_2H_AAD, calc.aad if calc.aad is not None else "-")
            w_cell(COL_2H_GI, int(calc.gi) if calc.gi is not None else "-")

            if not row_is_partial[start_row]:
                w_cell(COL_2H_MAD, calc.mad if calc.mad is not None else "-")
                w_cell(COL_2H_AD, round(calc.ash_dry, 2) if calc.ash_dry is not None else "-")
                w_cell(COL_2H_VAD, round(calc.vad, 2) if calc.vad is not None else "-")
                v_daf = getattr(calc, 'volatiles_daf', None)
                w_cell(COL_2H_VDAF, round(v_daf, 2) if v_daf is not None else "-")

        # =========================================================================
        # 5. ХЭСЭГ 2: 4 HOURLY
        # =========================================================================
        samples_4h = list(db.session.execute(
            select(Sample).where(
                Sample.received_date >= start_time,
                Sample.received_date <= end_time,
                Sample.client_name == 'CHPP',
                Sample.sample_type.in_(['4 hourly', '4 Hourly']),
            )
        ).scalars().all())

        for s in samples_4h:
            hour = s.received_date.hour
            code_upper = (s.sample_code or "").strip().upper()
            calc = s.get_calculations()

            base_row = None
            for (h_start, h_end), r_num in ROW_BUCKETS_4H.items():
                if h_start < h_end:
                    if h_start <= hour < h_end:
                        base_row = r_num
                        break
                else:
                    if hour >= h_start or hour < h_end:
                        base_row = r_num
                        break

            target_cols = None
            row_offset = 0

            for cf_key, conf in CF_OFFSET_MAP.items():
                if cf_key in code_upper:
                    target_cols = conf["cols"]
                    row_offset = conf["offset"]
                    break

            if base_row and target_cols:
                final_row = base_row + row_offset

                val_aad = calc.aad if calc.aad is not None else "-"
                cell_aad = ws.cell(row=final_row, column=target_cols['Aad'], value=val_aad)
                cell_aad.alignment = center_align
                cell_aad.font = font_reg

                val_fm = getattr(calc, 'fm', None)
                if val_fm is None:
                    val_fm = calc.mt
                final_fm = val_fm if val_fm is not None else "-"

                cell_fm = ws.cell(row=final_row, column=target_cols['FM'], value=final_fm)
                cell_fm.alignment = center_align
                cell_fm.font = font_reg

        # =========================================================================
        # 6. ХАДГАЛАХ
        # =========================================================================
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        filename = f"Hourly_Report_{report_date_str}_{report_time_str.replace(':', '')}.xlsx"

        sender_name = current_user.full_name or current_user.username
        sender_position = current_user.position or "Senior Chemist, Laboratory"
        sender_email = current_user.email or "lab@energyresources.mn"
        sender_phone = current_user.phone or ""

        phone_display = f"|Mobile: (976) {sender_phone}" if sender_phone else ""

        # =====================================================================
        # AI ANALYTICS — Anomaly detection, quality score, insights
        # =====================================================================
        all_period_samples = samples_2h + samples_4h
        ai_section = ""

        try:
            anomalies = detect_anomalies(all_period_samples)
            quality = calculate_quality_score(all_period_samples, anomalies)

            trend_codes = ["Mad", "Aad", "Vdaf"]
            trends = [analyze_trends(c, days=30, client_name="CHPP") for c in trend_codes]
            trends = [t for t in trends if t.confidence != "low"]

            insights = generate_insights(all_period_samples, anomalies, trends)

            ai_section = _build_ai_email_section(
                quality, anomalies, trends, insights, len(all_period_samples)
            )

        except Exception as ai_err:
            current_app.logger.warning("AI analytics failed (non-blocking): %s", ai_err)
            ai_section = ""

        email_subject = (
            f"Hourly Report {report_time_str}"
            + (f" | Quality: {quality.grade} ({quality.score}%)" if ai_section else "")
        )

        email_html = f"""
        <div style="font-family: Arial, sans-serif; font-size: 14px; color: #000000;">
            <p>Dear all,</p>
            <p>Please see the <strong>{report_time_str}</strong> hour sample results from the attachment.</p>
            {ai_section}
            <p>Regards,</p>
            <p>
                <b>{sender_name}</b><br>
                {sender_position}
            </p>
            <p>
                <b>ENERGY RESOURCES LLC</b><br>
                | Ukhaa Khudag Branch, Tsogttsetsii soum, Umnugobi province 46040 , MONGOLIA|<br>
                |Tel.: (976)7012 2279, 7013 2279 |Fax: (976) 11 322279 {phone_display}<br>
                |{sender_email} | <a href="http://www.mmc.mn/">http://www.mmc.mn/</a> |
            </p>
            <div style="border-top: 1px solid #000; margin-top: 10px; padding-top: 5px;">
                <span style="font-size: 11px; color: #555;">
                    This email is CONFIDENTIAL and is intended only for
                    the use of the person to whom it is addressed.<br>
                    Any distribution, copying or other use by anyone else
                    is strictly prohibited.<br>
                    If you have received this email in error, please telephone
                    or email us immediately and destroy this email.
                </span>
            </div>
        </div>
        """

        email_recipients = get_report_email_recipients()
        to_list = email_recipients['to']
        cc_list = email_recipients['cc']

        if not to_list:
            flash(_l("Имэйл хүлээн авагч тохируулагдаагүй байна. Тохиргооноос тохируулна уу."), "warning")
            return redirect(url_for('main.index'))

        msg = Message(
            subject=email_subject,
            recipients=to_list,
            cc=cc_list if cc_list else None,
            html=email_html
        )
        msg.attach(filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", final_output.read())
        mail.send(msg)

        sent_to = ", ".join(to_list)
        if cc_list:
            sent_to += f" (CC: {', '.join(cc_list)})"
        flash(_l("Амжилттай илгээгдлээ! → %(to)s") % {"to": sent_to}, "success")

    except (OSError, RuntimeError, ValueError):
        current_app.logger.exception("Error in send_hourly_report")
        flash(_l("Имэйл илгээхэд алдаа гарлаа."), "danger")

    return redirect(url_for('main.index'))


def _build_ai_email_section(quality, anomalies, trends, insights, sample_count):
    """AI analytics хэсгийг email HTML болгон бэлдэх."""
    if sample_count == 0 and not anomalies:
        return ""

    # ── Quality Score badge ──
    score = quality.score
    grade = quality.grade
    color = quality.color

    score_html = f"""
    <div style="margin: 20px 0; padding: 16px; background: #f8fafc; border-radius: 8px;
                border-left: 4px solid {color};">
        <table cellpadding="0" cellspacing="0" border="0" width="100%">
        <tr>
            <td width="80" style="text-align: center; vertical-align: middle;">
                <div style="width: 64px; height: 64px; border-radius: 50%;
                            background: {color}; color: white;
                            font-size: 24px; font-weight: bold;
                            line-height: 64px; text-align: center;
                            margin: 0 auto;">
                    {grade}
                </div>
            </td>
            <td style="padding-left: 16px; vertical-align: middle;">
                <div style="font-size: 18px; font-weight: bold; color: #1e293b;">
                    Quality Score: {score}/100
                </div>
                <div style="font-size: 12px; color: #64748b; margin-top: 4px;">
                    {quality.message}
                </div>
                <div style="font-size: 11px; color: #94a3b8; margin-top: 2px;">
                    Completeness: {quality.breakdown.get('completeness', 0):.0f}/30 |
                    Accuracy: {quality.breakdown.get('accuracy', 0):.0f}/30 |
                    Timeliness: {quality.breakdown.get('timeliness', 0):.0f}/20 |
                    Consistency: {quality.breakdown.get('consistency', 0):.0f}/20
                </div>
            </td>
        </tr>
        </table>
    </div>
    """

    # ── Anomalies table ──
    anomaly_html = ""
    if anomalies:
        critical = [a for a in anomalies if a.severity == "critical"]
        warnings = [a for a in anomalies if a.severity == "warning"]

        rows = ""
        for a in anomalies[:10]:
            sev_color = "#dc2626" if a.severity == "critical" else "#d97706"
            sev_bg = "#fef2f2" if a.severity == "critical" else "#fffbeb"
            sev_label = "CRITICAL" if a.severity == "critical" else "WARNING"
            spec = ANALYSIS_SPECS.get(a.analysis_code, {})
            name = spec.get("name", a.analysis_code)

            rows += f"""
            <tr style="background: {sev_bg};">
                <td style="padding: 6px 10px; border: 1px solid #e2e8f0;">
                    <span style="color: {sev_color}; font-weight: bold; font-size: 11px;">
                        {sev_label}
                    </span>
                </td>
                <td style="padding: 6px 10px; border: 1px solid #e2e8f0;">{a.sample_code}</td>
                <td style="padding: 6px 10px; border: 1px solid #e2e8f0;">{name}</td>
                <td style="padding: 6px 10px; border: 1px solid #e2e8f0; font-weight: bold;">
                    {a.value}
                </td>
                <td style="padding: 6px 10px; border: 1px solid #e2e8f0; color: #64748b; font-size: 12px;">
                    z={a.z_score:+.1f} | avg={a.historical_mean:.3f}
                </td>
                <td style="padding: 6px 10px; border: 1px solid #e2e8f0; font-size: 11px; color: #475569;">
                    {a.recommendation}
                </td>
            </tr>"""

        title_color = "#dc2626" if critical else "#d97706"
        anomaly_html = f"""
        <div style="margin: 16px 0;">
            <div style="font-size: 14px; font-weight: bold; color: {title_color}; margin-bottom: 8px;">
                Anomaly Detection ({len(critical)} critical, {len(warnings)} warning)
            </div>
            <table cellpadding="0" cellspacing="0" border="0" width="100%"
                   style="border-collapse: collapse; font-size: 12px;">
                <tr style="background: #f1f5f9;">
                    <th style="padding: 6px 10px; border: 1px solid #e2e8f0; text-align: left;">Level</th>
                    <th style="padding: 6px 10px; border: 1px solid #e2e8f0; text-align: left;">Sample</th>
                    <th style="padding: 6px 10px; border: 1px solid #e2e8f0; text-align: left;">Analysis</th>
                    <th style="padding: 6px 10px; border: 1px solid #e2e8f0; text-align: left;">Value</th>
                    <th style="padding: 6px 10px; border: 1px solid #e2e8f0; text-align: left;">Stats</th>
                    <th style="padding: 6px 10px; border: 1px solid #e2e8f0; text-align: left;">Action</th>
                </tr>
                {rows}
            </table>
        </div>
        """

    # ── Trends ──
    trend_html = ""
    if trends:
        trend_items = ""
        for t in trends:
            if t.direction == "stable":
                icon, t_color = "→", "#22c55e"
            elif t.direction == "increasing":
                icon, t_color = "↑", "#f59e0b"
            else:
                icon, t_color = "↓", "#3b82f6"

            t_name = ANALYSIS_SPECS.get(t.analysis_code, {}).get("name", t.analysis_code)
            trend_items += f"""
            <div style="display: inline-block; margin: 4px 8px 4px 0; padding: 6px 12px;
                        background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px;">
                <span style="color: {t_color}; font-weight: bold; font-size: 16px;">{icon}</span>
                <span style="font-size: 12px; color: #334155;">
                    {t_name}: {t.change_pct:+.1f}%
                </span>
                <span style="font-size: 10px; color: #94a3b8;">(R²={t.r_squared:.2f})</span>
            </div>"""

        trend_html = f"""
        <div style="margin: 16px 0;">
            <div style="font-size: 14px; font-weight: bold; color: #334155; margin-bottom: 8px;">
                30-Day Trends
            </div>
            {trend_items}
        </div>
        """

    # ── Insights ──
    insight_html = ""
    if insights:
        items = "".join(
            f'<li style="margin-bottom: 4px; color: #475569;">{i}</li>'
            for i in insights
        )
        insight_html = f"""
        <div style="margin: 16px 0; padding: 12px 16px; background: #eff6ff;
                    border-radius: 8px; border: 1px solid #bfdbfe;">
            <div style="font-size: 13px; font-weight: bold; color: #1e40af; margin-bottom: 6px;">
                AI Insights
            </div>
            <ul style="margin: 0; padding-left: 20px; font-size: 12px;">
                {items}
            </ul>
        </div>
        """

    return f"""
    <div style="margin: 20px 0; border-top: 2px solid #e2e8f0; padding-top: 16px;">
        <div style="font-size: 11px; color: #94a3b8; text-transform: uppercase;
                    letter-spacing: 1px; margin-bottom: 12px;">
            Laboratory Intelligence Report ({sample_count} samples analyzed)
        </div>
        {score_html}
        {anomaly_html}
        {trend_html}
        {insight_html}
    </div>
    """
