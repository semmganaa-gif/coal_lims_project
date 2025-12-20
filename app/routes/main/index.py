# app/routes/main/index.py
# -*- coding: utf-8 -*-
"""
Нүүр хуудас - Дээж бүртгэх (Index/Registration) routes
"""

# 1. Standard Library Imports (Python-ы үндсэн сангууд)
import os
from io import BytesIO
from datetime import timedelta

# 2. Third-Party Imports (Гараас суулгасан сангууд)
from flask import render_template, flash, redirect, url_for, request, jsonify, current_app
from flask_login import login_required, current_user
from flask_mail import Message
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# 3. Local Application Imports (Таны төслийн файлууд)
from app import db, mail
from app.models import Sample, SystemSetting
from app.forms import AddSampleForm

# Current Blueprint & Local Helpers
from . import main_bp
from .helpers import get_12h_shift_code, get_quarter_code

# Utils
from app.utils.datetime import now_local
from app.utils.analysis_assignment import assign_analyses_to_sample


def get_report_email_recipients():
    """
    Тайлан илгээх имэйл хаягуудыг SystemSetting-ээс авах

    Returns:
        dict: {'to': [...], 'cc': [...]}
    """
    result = {'to': [], 'cc': []}

    # TO хаягууд
    to_setting = SystemSetting.query.filter_by(
        category='email',
        key='report_recipients_to',
        is_active=True
    ).first()
    if to_setting and to_setting.value:
        result['to'] = [e.strip() for e in to_setting.value.split(',') if e.strip()]

    # CC хаягууд
    cc_setting = SystemSetting.query.filter_by(
        category='email',
        key='report_recipients_cc',
        is_active=True
    ).first()
    if cc_setting and cc_setting.value:
        result['cc'] = [e.strip() for e in cc_setting.value.split(',') if e.strip()]

    return result
from app.utils.sorting import custom_sample_sort_key
from app.utils.database import safe_commit
from app.utils.settings import get_sample_type_choices_map, get_unit_abbreviations

# Monitoring
from app.monitoring import track_sample

# Constants
from app.constants import (
    ALL_12H_SAMPLES,
    CONSTANT_12H_SAMPLES,
    COM_PRIMARY_PRODUCTS,
    COM_SECONDARY_MAP,
    WTL_SAMPLE_NAMES_19,
    WTL_SAMPLE_NAMES_70,
    WTL_SAMPLE_NAMES_6,
    WTL_SAMPLE_NAMES_2,
    WTL_SIZE_NAMES,
    WTL_FL_NAMES,
)

def register_routes(bp):
    """Route-уудыг өгөгдсөн blueprint дээр бүртгэх"""

    # =====================================================================
    # 1. НҮҮР ХУУДАС / ДЭЭЖ БҮРТГЭХ
    # =====================================================================
    @bp.route("/", methods=["GET", "POST"])
    @bp.route("/index", methods=["GET", "POST"])
    @login_required
    def index():
        """
        - Зүүн тал: жагсаалт (DataTables)
        - Баруун тал: 'Шинэ дээж бүртгэх' form
        """
        from app.models import Sample

        form = AddSampleForm()
        sample_type_choices_map = get_sample_type_choices_map()  # ✅ DB

        # client_name choices
        form.client_name.choices = [
            ("CHPP", "CHPP"),
            ("UHG-Geo", "UHG-Geo"),
            ("BN-Geo", "BN-Geo"),
            ("QC", "QC"),
            ("Proc", "Proc"),
            ("WTL", "WTL"),
            ("LAB", "LAB"),
        ]

        # ---------- Dynamic sample_type + validators ----------
        if request.method == "POST":
            selected_client = request.form.get("client_name")
            form.sample_type.choices = [
                (v, v) for v in sample_type_choices_map.get(selected_client, [])
            ]
            if selected_client == "WTL" and request.form.get("sample_type") in ["MG", "Test"]:
                from wtforms.validators import DataRequired
                form.sample_code.validators = [DataRequired(message="Sample name заавал оруулна уу.")]
            else:
                from wtforms.validators import Optional
                form.sample_code.validators = [Optional()]
        else:
            form.sample_type.choices = []

        # ---------- Үндсэн submit ----------
        if form.validate_on_submit():
            if current_user.role not in ["prep", "admin"]:
                flash("Дээж бүртгэх эрх танд байхгүй.", "danger")
                return redirect(url_for("main.index"))

            raw_submitted_codes = request.form.getlist("sample_codes")
            list_type = request.form.get("list_type")  # chpp_2h / chpp_4h / ...
            client_name = form.client_name.data
            sample_type = form.sample_type.data
            sample_date_obj = form.sample_date.data

            successful_samples, failed_samples = [], []
            count = 0

            # --- 1) ОЛОН ДЭЭЖ БҮРТГЭХ (CHPP, UHG/BN/QC/Proc/LAB(simple)) ---
            if raw_submitted_codes and list_type:
                requires_weight = (
                    list_type == "chpp_2h"
                    or list_type == "multi_gen"
                    or list_type == "chpp_com"
                )

                # Жингийн мэдээллийг Map болгож авна (Code -> Weight)
                raw_weights = request.form.getlist("weights") if requires_weight else []
                code_weight_map = {}
                for idx, c in enumerate(raw_submitted_codes):
                    w = raw_weights[idx] if idx < len(raw_weights) else None
                    code_weight_map[c] = w

                # ✅ ЛОГИК ЗАСВАР: Ирсэн кодуудыг "Хатуу дүрмээр" эрэмбэлнэ.
                sorted_codes = sorted(raw_submitted_codes, key=custom_sample_sort_key)

                # ✅ Бүх үйлдлийг нэг transaction-д багтаана (atomicity)
                try:
                    for code in sorted_codes:
                        if not code:
                            continue

                        # Жин шалгах
                        weight, is_valid = None, True
                        if requires_weight:
                            weight_str = code_weight_map.get(code)
                            if weight_str:
                                try:
                                    weight = float(weight_str)
                                    # Жингийн утгын хязгаар шалгах
                                    from app.constants import MIN_SAMPLE_WEIGHT, MAX_SAMPLE_WEIGHT
                                    if weight <= MIN_SAMPLE_WEIGHT:
                                        failed_samples.append(f'{code} (жин хэт бага байна: {weight}г)')
                                        is_valid = False
                                    elif weight > MAX_SAMPLE_WEIGHT:
                                        failed_samples.append(f'{code} (жин хэт том байна: {weight}г, max {MAX_SAMPLE_WEIGHT}г)')
                                        is_valid = False
                                except ValueError:
                                    failed_samples.append(f'{code} (жин: "{weight_str}" буруу)')
                                    is_valid = False
                            else:
                                failed_samples.append(f"{code} (жин оруулаагүй)")
                                is_valid = False

                        if not is_valid:
                            continue

                        # Дээж үүсгэх
                        sample = Sample(
                            sample_code=code,
                            weight=weight,
                            user_id=current_user.id,
                            client_name=client_name,
                            sample_type=sample_type,
                            sample_condition=form.sample_condition.data,
                            sample_date=sample_date_obj,
                            return_sample=form.return_sample.data,
                            retention_date=(now_local() + timedelta(days=int(form.retention_period.data or 7))).date(),
                            delivered_by=form.delivered_by.data,
                            prepared_date=form.prepared_date.data,
                            prepared_by=form.prepared_by.data,
                            notes=form.notes.data,
                            location=form.location.data if list_type == "multi_gen" else None,
                            product=form.product.data if list_type == "multi_gen" and (client_name == "QC" or client_name == "Proc") else None,
                            hourly_system=list_type.replace("chpp_", "") if "chpp" in list_type else None,
                            analyses_to_perform="[]"
                        )

                        # Шинжилгээ оноох
                        # ✅ no_autoflush: Query хийхэд sample-г flush хийхгүй байх
                        with db.session.no_autoflush:
                            assign_analyses_to_sample(sample)

                        db.session.add(sample)
                        successful_samples.append(code)
                        count += 1

                except Exception as e:
                    # Loop-н дундаас алдаа гарвал бүх partial changes-ийг rollback хийнэ
                    db.session.rollback()
                    current_app.logger.error(f"Error during sample registration loop: {e}")
                    flash(f"Дээж бүртгэх үед алдаа гарлаа: {str(e)}", "danger")
                    # Continue to render template with errors below

                # ✅ Сайжруулсан error handling - давхардлыг арилгасан
                if count > 0:
                    if not safe_commit(
                        f"{count} ш дээж амжилттай бүртгэгдлээ!",
                        "БҮРТГЭЛ АМЖИЛТГҮЙ: Дээжний код давхардсан байна."
                    ):
                        # ✅ Бүх template variable дамжуулах
                        return render_template(
                            "index.html",
                            title="Нүүр хуудас",
                            form=form,
                            active_tab="add-pane",
                            sample_type_map=get_sample_type_choices_map(),
                            all_12h_samples=ALL_12H_SAMPLES,
                            constant_12h_samples=CONSTANT_12H_SAMPLES,
                            com_primary_products=COM_PRIMARY_PRODUCTS,
                            com_secondary_map=COM_SECONDARY_MAP,
                            unit_abbreviations=get_unit_abbreviations(),
                        )

                if failed_samples:
                    flash(f'Анхаар: Дараах дээжнүүд бүртгэгдсэнгүй: {", ".join(failed_samples)}', "warning")

                # ✅ Prometheus metrics: Дээж бүртгэлийг track хийх
                for _ in range(count):
                    track_sample(client=client_name, sample_type=sample_type)

                return redirect(url_for("main.index", active_tab="add-pane"))

            # --- 2) WTL (WTL/Size/FL) – автоматаар олон нэр үүсгэх ---
            elif not list_type and client_name == "WTL" and sample_type in ["WTL", "Size", "FL"]:
                lab_number = form.lab_number.data
                if not lab_number:
                    flash("WTL-ийн хувьд Лаб дугаар заавал оруулна уу.", "danger")
                else:
                    all_wtl_names = []
                    if sample_type == "WTL":
                        all_wtl_names = (WTL_SAMPLE_NAMES_19 + WTL_SAMPLE_NAMES_70 + WTL_SAMPLE_NAMES_6 + WTL_SAMPLE_NAMES_2)
                    elif sample_type == "Size":
                        all_wtl_names = WTL_SIZE_NAMES
                    elif sample_type == "FL":
                        all_wtl_names = WTL_FL_NAMES

                    count = 0
                    # ✅ try блокийг устгасан - safe_commit ашиглаж байгаа учир
                    for name in all_wtl_names:
                        final_sample_code = f"{lab_number}_{name}"

                        sample = Sample(
                            sample_code=final_sample_code,
                            user_id=current_user.id,
                            client_name=client_name,
                            sample_type=sample_type,
                            sample_condition=form.sample_condition.data,
                            sample_date=sample_date_obj,
                            return_sample=form.return_sample.data,
                            retention_date=(now_local() + timedelta(days=int(form.retention_period.data or 7))).date(),
                            delivered_by=form.delivered_by.data,
                            prepared_date=form.prepared_date.data,
                            prepared_by=form.prepared_by.data,
                            notes=form.notes.data,
                            weight=None,
                            analyses_to_perform="[]"
                        )

                        # ✅ no_autoflush: Query хийхэд sample-г flush хийхгүй байх
                        with db.session.no_autoflush:
                            assign_analyses_to_sample(sample)
                        db.session.add(sample)
                        count += 1

                    # ✅ Сайжруулсан error handling
                    if count > 0:
                        if not safe_commit(
                            f"{count} ш {sample_type} дээж амжилттай бүртгэгдлээ!",
                            "БҮРТГЭЛ АМЖИЛТГҮЙ: Дээжний код давхардсан байна."
                        ):
                            # ✅ Бүх template variable дамжуулах
                            return render_template(
                                "index.html",
                                title="Нүүр хуудас",
                                form=form,
                                active_tab="add-pane",
                                sample_type_map=get_sample_type_choices_map(),
                                all_12h_samples=ALL_12H_SAMPLES,
                                constant_12h_samples=CONSTANT_12H_SAMPLES,
                                com_primary_products=COM_PRIMARY_PRODUCTS,
                                com_secondary_map=COM_SECONDARY_MAP,
                                unit_abbreviations=get_unit_abbreviations(),
                            )

                    return redirect(url_for("main.index", active_tab="add-pane"))

            # --- 3) LAB — автоматаар нэр үүсгэх ---
            elif not list_type and client_name == "LAB":
                formatted_date = sample_date_obj.strftime("%Y%m%d")
                shift_code = get_12h_shift_code(now_local())

                if sample_type == "CM":
                    # Идэвхтэй CM стандартын нэрийг авах
                    from app.models import ControlStandard
                    active_cm = ControlStandard.query.filter_by(is_active=True).first()
                    cm_name = active_cm.name if active_cm else "CM"
                    # CM стандарт нэр аль хэдийн улирал агуулсан (жнь: CM-2025-Q4)
                    # Тиймээс quarter_code нэмэхгүй - давхардахаас сэргийлж
                    final_sample_code = f"{cm_name}_{formatted_date}{shift_code}"
                elif sample_type == "GBW":
                    # Идэвхтэй GBW стандартын нэрийг авах
                    from app.models import GbwStandard
                    active_gbw = GbwStandard.query.filter_by(is_active=True).first()
                    gbw_name = active_gbw.name if active_gbw else "GBW"
                    final_sample_code = f"{gbw_name}_{formatted_date}{shift_code}"
                elif sample_type == "Test":
                    final_sample_code = f"Test_{formatted_date}{shift_code}"
                else:
                    final_sample_code = f"LAB_UNKNOWN_{formatted_date}"

                sample = Sample(
                    sample_code=final_sample_code,
                    user_id=current_user.id,
                    client_name=client_name,
                    sample_type=sample_type,
                    sample_condition=form.sample_condition.data,
                    sample_date=sample_date_obj,
                    return_sample=form.return_sample.data,
                    delivered_by=form.delivered_by.data,
                    prepared_date=form.prepared_date.data,
                    prepared_by=form.prepared_by.data,
                    notes=form.notes.data,
                    weight=None,
                    analyses_to_perform="[]"
                )

                # ✅ no_autoflush: Query хийхэд sample-г flush хийхгүй байх
                with db.session.no_autoflush:
                    assign_analyses_to_sample(sample)
                db.session.add(sample)
                # ✅ Сайжруулсан error handling
                safe_commit(
                    f"Дээж амжилттай бүртгэгдлээ: {final_sample_code}",
                    f'БҮРТГЭЛ АМЖИЛТГҮЙ: "{final_sample_code}" нэртэй дээж аль хэдийн бүртгэлтэй байна.'
                )
                return redirect(url_for("main.index", active_tab="add-pane"))

            # --- 4) WTL – MG/Test (гар аргаар sample_code) ---
            elif not list_type and client_name == "WTL" and sample_type in ["MG", "Test"]:
                if not form.sample_code.data:
                    flash("WTL-ийн энэ төрлийн хувьд Sample name заавал оруулна уу.", "danger")
                else:
                    final_sample_code = form.sample_code.data
                    sample = Sample(
                        sample_code=final_sample_code,
                        user_id=current_user.id,
                        client_name=client_name,
                        sample_type=sample_type,
                        sample_condition=form.sample_condition.data,
                        sample_date=sample_date_obj,
                        return_sample=form.return_sample.data,
                        delivered_by=form.delivered_by.data,
                        prepared_date=form.prepared_date.data,
                        prepared_by=form.prepared_by.data,
                        notes=form.notes.data,
                        analyses_to_perform="[]"
                    )

                    # ✅ no_autoflush: Query хийхэд sample-г flush хийхгүй байх
                    with db.session.no_autoflush:
                        assign_analyses_to_sample(sample)
                    db.session.add(sample)
                    # ✅ Сайжруулсан error handling
                    safe_commit(
                        "Шинэ дээж амжилттай бүртгэгдлээ!",
                        f'БҮРТГЭЛ АМЖИЛТГҮЙ: "{final_sample_code}" нэртэй дээж аль хэдийн бүртгэлтэй байна.'
                    )
                    return redirect(url_for("main.index", active_tab="add-pane"))

            else:
                flash("Форм бүрэн гүйцэд бөглөгдөөгүй эсвэл буруу байна.", "danger")

        active_tab = "add-pane" if form.errors else request.args.get("active_tab", "list-pane")

        return render_template(
            "index.html",
            title="Нүүр хуудас",
            form=form,
            active_tab=active_tab,
            sample_type_map=get_sample_type_choices_map(),
            all_12h_samples=ALL_12H_SAMPLES,
            constant_12h_samples=CONSTANT_12H_SAMPLES,
            com_primary_products=COM_PRIMARY_PRODUCTS,
            com_secondary_map=COM_SECONDARY_MAP,
            unit_abbreviations=get_unit_abbreviations(),
        )

    # =====================================================================
    # 2. LIVE PREVIEW (AJAX)
    # =====================================================================
    @bp.route("/preview-analyses", methods=["POST"])
    @login_required
    def preview_sample_analyses():
        """
        Frontend-ээс (JS) дээжний нэр ирэхэд ямар анализ хийгдэхийг урьдчилан харуулна.
        """
        data = request.json
        sample_names = data.get("sample_names", [])
        client_name = data.get("client_name")
        sample_type = data.get("sample_type")

        if not all([sample_names, client_name, sample_type]):
            return jsonify({"error": "Missing data"}), 400

        results = {}

        # ✅ САЙЖРУУЛСАН: Fake объект үүсгэхгүй, шууд параметр дамжуулах
        for name in sample_names:
            assigned_list = assign_analyses_to_sample(
                client_name=client_name,
                sample_type=sample_type,
                sample_code=name
            )
            results[name] = assigned_list

        return jsonify(results)

# app/routes/main/index.py

@main_bp.route("/send-hourly-report")
def send_hourly_report():
    try:
        current_app.logger.debug("HOURLY REPORT STARTED (FIXED POSITIONING)")

        # =========================================================================
        # 1. ТОХИРГОО
        # =========================================================================

        # --- 2 HOURLY MAPPING (Суурь мөрүүд) ---
        PF_MAPPING = { 'PF211': 20, 'PF221': 33, 'PF231': 46 }
        HCC_KEYWORDS = ['UHG MV', 'UHG HV', 'BN HV', 'BN SSCC']
        MIDD_KEYWORDS = ['BN MASHCC', 'BN MIDD', 'UHG MASHCC', 'UHG MIDD', 'MASHCC_2']

        # D/N дугаарлалт -> Мөрийн шилжилт (Offset)
        # D1 нь тухайн группын эхний мөр, N6 нь 12 дахь мөр байна.
        SUFFIX_OFFSET_MAP = {
            'D1': 0, 'D2': 1, 'D3': 2, 'D4': 3, 'D5': 4, 'D6': 5,
            'N1': 6, 'N2': 7, 'N3': 8, 'N4': 9, 'N5': 10, 'N6': 11
        }

        # 2H Баганын дугаарууд
        COL_2H_NAME = 1   # A: Sample ID
        # COL_2H_TIME = 2 # B: Time (ОРОЛДОХГҮЙ)
        COL_2H_MT   = 3   # C
        COL_2H_MAD  = 4   # D
        COL_2H_AAD  = 5   # E
        COL_2H_AD   = 6   # F
        COL_2H_VAD  = 7   # G
        COL_2H_VDAF = 8   # H
        COL_2H_GI   = 15  # O

        # --- 4 HOURLY CONFIG ---
        ROW_BUCKETS_4H = {
            (8, 12):  87,  # 10:00
            (12, 16): 91,  # 14:00
            (16, 20): 95,  # 18:00
            (20, 24): 99,  # 22:00
            (0, 4):   103, # 02:00
            (4, 8):   107  # 06:00
        }

        CF_OFFSET_MAP = {
            # MOD I
            "CF501":       { "cols": {"Aad": 3, "FM": 5}, "offset": 0 },
            "CF502":       { "cols": {"Aad": 3, "FM": 5}, "offset": 1 },
            "CF601":       { "cols": {"Aad": 3, "FM": 5}, "offset": 2 },
            "CF602":       { "cols": {"Aad": 3, "FM": 5}, "offset": 2 }, # 601/602
            # MOD II
            "CF521":       { "cols": {"Aad": 11, "FM": 13}, "offset": 0 },
            "CF522":       { "cols": {"Aad": 11, "FM": 13}, "offset": 1 },
            "CF621":       { "cols": {"Aad": 11, "FM": 13}, "offset": 2 },
            "CF622":       { "cols": {"Aad": 11, "FM": 13}, "offset": 2 },
            # MOD III
            "CF541":       { "cols": {"Aad": 19, "FM": 21}, "offset": 0 },
            "CF542":       { "cols": {"Aad": 19, "FM": 21}, "offset": 1 },
            "CF641":       { "cols": {"Aad": 19, "FM": 21}, "offset": 2 },
            "CF642":       { "cols": {"Aad": 19, "FM": 21}, "offset": 2 },
        }

        # =========================================================================
        # 2. ЦАГ ХУГАЦАА
        # =========================================================================
        current_time = now_local()
        # Тайлангийн цаг (Subject-д зориулж)
        calc_time = current_time - timedelta(hours=2)
        report_hour = (calc_time.hour // 2) * 2
        report_dt = calc_time.replace(hour=report_hour, minute=0, second=0, microsecond=0)

        report_time_str = report_dt.strftime('%H:%M')
        report_date_str = report_dt.strftime('%Y.%m.%d')

        # Data Start Time (08:00 boundary)
        if report_dt.hour < 8:
            start_time = (report_dt - timedelta(days=1)).replace(hour=8, minute=0, second=0, microsecond=0)
        else:
            start_time = report_dt.replace(hour=8, minute=0, second=0, microsecond=0)

        end_time = current_time

        # =========================================================================
        # 3. EXCEL НЭЭХ
        # =========================================================================
        template_path = os.path.join(current_app.root_path, 'static', 'hourly_template.xlsx')
        if not os.path.exists(template_path):
            if os.path.exists(template_path + ".xlsx"): template_path += ".xlsx"
            else:
                flash("Загвар файл олдсонгүй!", "danger")
                return redirect(url_for('main.index'))

        with open(template_path, "rb") as f:
            output = BytesIO(f.read())

        wb = load_workbook(output)
        ws = wb.active

        center_align = Alignment(horizontal='center', vertical='center')
        font_reg = Font(name='Times New Roman', size=12)
        font_bold = Font(name='Times New Roman', size=12, bold=False)

        # -------------------------------------------------------------
        # HEADER & COUNTER
        # -------------------------------------------------------------
        setting_key = f"report_counter_{start_time.year}"
        last_update_key = f"last_update_{start_time.year}"

        setting_count = SystemSetting.query.filter_by(category='report_config', key=setting_key).first()
        setting_last = SystemSetting.query.filter_by(category='report_config', key=last_update_key).first()

        current_count = int(setting_count.value) if setting_count else 0
        last_updated_date = setting_last.value if setting_last else ""
        today_str = start_time.strftime('%Y-%m-%d')

        if report_time_str == "07:00" and last_updated_date != today_str:
            current_count += 1
            if not setting_count:
                db.session.add(SystemSetting(category='report_config', key=setting_key, value=str(current_count)))
            else:
                setting_count.value = str(current_count)
            if not setting_last:
                db.session.add(SystemSetting(category='report_config', key=last_update_key, value=today_str))
            else:
                setting_last.value = today_str
            db.session.commit()

        display_count = current_count if current_count > 0 else 1

        ws['E10'] = f"{start_time.year}_{display_count:03d}"
        ws['E10'].font = font_reg
        ws['E10'].alignment = Alignment(horizontal='left', vertical='center')

        ws['E11'] = start_time.strftime('%d-%b-%Y')
        ws['E11'].font = font_reg
        ws['E11'].alignment = Alignment(horizontal='left', vertical='center')

        ws['O3'] = f"№ {start_time.strftime('%y-%m-%d')}"
        ws['O3'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['O3'].alignment = Alignment(horizontal='right', vertical='center')

        ws['C4'] = start_time.strftime('%Y.%m.%d')
        ws['C4'].font = font_reg

        # =========================================================================
        # 4. ХЭСЭГ 1: 2 HOURLY ДЭЭЖҮҮД (FIXED POSITIONING)
        # =========================================================================
        samples_2h = Sample.query.filter(
            Sample.received_date >= start_time,
            Sample.received_date <= end_time,
            Sample.client_name == 'CHPP',
            Sample.sample_type.in_(['2 hourly', '2 Hourly'])
        ).all()

        # Энэ dictionary нь мөр тус бүрт "Partial" (nr бичих) эсэхийг хадгална
        row_is_partial = {20: True, 33: True, 46: True, 59: False, 72: False}

        for s in samples_2h:
            code_upper = (s.sample_code or "").strip().upper()
            calc = s.get_calculations()

            # 1. СУУРЬ МӨРӨӨ ОЛОХ (PF211, UHG MV...)
            start_row = None
            for pf_key, r_num in PF_MAPPING.items():
                if pf_key in code_upper: start_row = r_num; break

            if not start_row:
                for k in HCC_KEYWORDS:
                    if k in code_upper: start_row = 59; break
                if not start_row and "UHG MASHCC" in code_upper and "MASHCC_2" not in code_upper: start_row = 59

            if not start_row:
                for k in MIDD_KEYWORDS:
                    if k in code_upper: start_row = 72; break

            # Олдохгүй бол UHG MV рүү
            if not start_row: start_row = 59

            # 2. ШИЛЖИЛТЭЭ ОЛОХ (D1, D2...)
            # Дээжийн нэрний сүүлийн хэсгийг шалгах (Жнь: ..._D1)
            row_offset = 0

            # SUFFIX_OFFSET_MAP-ээс хайх (D1, N6...)
            # Sample code-ийн төгсгөлд байгаа эсэхийг шалгана, эсвэл "_" дараа байгааг
            for suffix, offset in SUFFIX_OFFSET_MAP.items():
                # Төгсгөлд нь "_D1" эсвэл зүгээр "D1" байгаа эсэх
                # Хамгийн найдвартай нь "_D1" гэж хайх
                if f"_{suffix}" in code_upper or code_upper.endswith(suffix):
                    row_offset = offset
                    break

            # Хэрэв D/N дугаар олдоогүй бол яах вэ?
            # Эрсдэлтэй тул эхний мөрөнд бичих үү, эсвэл алгасах уу?
            # Одоохондоо Offset=0 (Эхний мөр) гэж үзье.

            final_row = start_row + row_offset

            # 3. БИЧИХ (LIMS Нэр, Тооцоо)
            def w_cell(c, v, f=font_reg):
                cell = ws.cell(row=final_row, column=c, value=v)
                cell.alignment = center_align
                cell.font = f

            # A багана: Жинхэнэ нэр
            w_cell(COL_2H_NAME, s.sample_code, font_bold)

            # B багана: БИЧИХГҮЙ (Загвар дээрх цагийг эвдэхгүй)

            # Үр дүнгүүд
            w_cell(COL_2H_MT, calc.mt if calc.mt is not None else "-")
            w_cell(COL_2H_AAD, calc.aad if calc.aad is not None else "-")
            w_cell(COL_2H_GI, int(calc.gi) if calc.gi is not None else "-")

            if not row_is_partial[start_row]:
                w_cell(COL_2H_MAD, calc.mad if calc.mad is not None else "-")
                w_cell(COL_2H_AD, round(calc.ash_dry, 2) if calc.ash_dry is not None else "-")
                w_cell(COL_2H_VAD, round(calc.vad, 2) if calc.vad is not None else "-")
                v_daf = getattr(calc, 'volatiles_daf', None)
                w_cell(COL_2H_VDAF, round(v_daf, 2) if v_daf is not None else "-")

        # =========================================================================
        # 5. ХЭСЭГ 2: 4 HOURLY
        # =========================================================================
        samples_4h = Sample.query.filter(
            Sample.received_date >= start_time,
            Sample.received_date <= end_time,
            Sample.client_name == 'CHPP',
            Sample.sample_type.in_(['4 hourly', '4 Hourly'])
        ).all()

        for s in samples_4h:
            hour = s.received_date.hour
            code_upper = (s.sample_code or "").strip().upper()
            calc = s.get_calculations()

            # 1. Мөр олох
            base_row = None
            for (h_start, h_end), r_num in ROW_BUCKETS_4H.items():
                if h_start < h_end:
                    if h_start <= hour < h_end: base_row = r_num; break
                else:
                    if hour >= h_start or hour < h_end: base_row = r_num; break

            # 2. Багана олох
            target_cols = None
            row_offset = 0

            for cf_key, conf in CF_OFFSET_MAP.items():
                if cf_key in code_upper:
                    target_cols = conf["cols"]
                    row_offset = conf["offset"]
                    break

            # 3. Бичих
            if base_row and target_cols:
                final_row = base_row + row_offset

                # A багана: Жинхэнэ нэр (4H дээр бичих үү? Та "үгүй" гэсэн байх. Хэрэв бичих бол доорхыг нээнэ үү)
                # cell_name = ws.cell(row=final_row, column=1, value=s.sample_code)
                # cell_name.font = font_bold
                # cell_name.alignment = center_align

                val_aad = calc.aad if calc.aad is not None else "-"
                cell_aad = ws.cell(row=final_row, column=target_cols['Aad'], value=val_aad)
                cell_aad.alignment = center_align
                cell_aad.font = font_reg

                val_fm = getattr(calc, 'fm', None)
                if val_fm is None: val_fm = calc.mt
                final_fm = val_fm if val_fm is not None else "-"

                cell_fm = ws.cell(row=final_row, column=target_cols['FM'], value=final_fm)
                cell_fm.alignment = center_align
                cell_fm.font = font_reg

        # =========================================================================
        # 6. ХАДГАЛАХ
        # =========================================================================
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        filename = f"Hourly_Report_{report_date_str}_{report_time_str.replace(':', '')}.xlsx"
        email_subject = f"Hourly Report {report_time_str}"

        # Нэвтэрсэн хэрэглэгчийн мэдээлэл (илгээж буй ахлах химич)
        sender_name = current_user.full_name or current_user.username
        sender_position = current_user.position or "Senior Chemist, Laboratory"
        sender_email = current_user.email or "lab@energyresources.mn"
        sender_phone = current_user.phone or ""

        # Phone форматлах
        phone_display = f"|Mobile: (976) {sender_phone}" if sender_phone else ""

        email_html = f"""
        <div style="font-family: Arial, sans-serif; font-size: 14px; color: #000000;">
            <p>Dear all,</p>
            <p>Please see the <strong>{report_time_str}</strong> hour sample results from the attachment.</p>
            <br>
            <p>Regards,</p>
            <p>
                <b>{sender_name}</b><br>
                {sender_position}
            </p>
            <p>
                <b>ENERGY RESOURCES LLC</b><br>
                | Ukhaa Khudag Branch, Tsogttsetsii soum, Umnugobi province 46040 , MONGOLIA|<br>
                |Tel.: (976)7012 2279, 7013 2279 |Fax: (976) 11 322279 {phone_display}<br>
                |{sender_email} | <a href="http://www.mmc.mn/">http://www.mmc.mn/</a> |
            </p>
            <div style="border-top: 1px solid #000; margin-top: 10px; padding-top: 5px;">
                <span style="font-size: 11px; color: #555;">
                    This email is CONFIDENTIAL and is intended only for the use of the person to whom it is addressed.<br>
                    Any distribution, copying or other use by anyone else is strictly prohibited.<br>
                    If you have received this email in error, please telephone or email us immediately and destroy this email.
                </span>
            </div>
        </div>
        """

        # Имэйл хүлээн авагчдыг SystemSetting-ээс авах
        email_recipients = get_report_email_recipients()
        to_list = email_recipients['to']
        cc_list = email_recipients['cc']

        # Хэрэв TO хаяг байхгүй бол алдаа
        if not to_list:
            flash("Имэйл хүлээн авагч тохируулагдаагүй байна. Тохиргоо хэсгээс оруулна уу.", "warning")
            return redirect(url_for('main.index'))

        msg = Message(
            subject=email_subject,
            recipients=to_list,
            cc=cc_list if cc_list else None,
            html=email_html
        )
        msg.attach(filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", final_output.read())
        mail.send(msg)

        # Илгээсэн хаягуудыг харуулах
        sent_to = ", ".join(to_list)
        if cc_list:
            sent_to += f" (CC: {', '.join(cc_list)})"
        flash(f"Амжилттай илгээгдлээ! → {sent_to}", "success")

    except Exception as e:
        current_app.logger.exception("Error in send_hourly_report")
        flash(f"Алдаа: {str(e)}", "danger")

    return redirect(url_for('main.index'))
