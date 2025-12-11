# app/utils/exports.py
# -*- coding: utf-8 -*-
"""
Export Utility - Excel/PDF export functions

ISO 17025 дагуу өгөгдөл экспорт хийх функцүүд.
"""

import logging
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
from flask import send_file

logger = logging.getLogger(__name__)


def export_to_excel(
    data: List[Dict[str, Any]],
    columns: List[Dict[str, str]],
    filename: str = "export.xlsx",
    sheet_name: str = "Data"
) -> BytesIO:
    """
    Өгөгдлийг Excel файл болгох

    Args:
        data: Өгөгдлийн жагсаалт [{"col1": val1, ...}, ...]
        columns: Баганы тодорхойлолт [{"key": "col1", "label": "Багана 1"}, ...]
        filename: Файлын нэр
        sheet_name: Sheet нэр

    Returns:
        BytesIO Excel файл
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    # DataFrame үүсгэх
    df_data = []
    for row in data:
        df_row = {}
        for col in columns:
            df_row[col['label']] = row.get(col['key'], '')
        df_data.append(df_row)

    df = pd.DataFrame(df_data)

    # Workbook үүсгэх
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Header бичих
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Data бичих
    for row_idx, row in enumerate(data, 2):
        for col_idx, col in enumerate(columns, 1):
            value = row.get(col['key'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Баганы өргөн автоматаар тохируулах
    for col_idx, col in enumerate(columns, 1):
        max_length = len(col['label'])
        for row in data:
            cell_value = str(row.get(col['key'], ''))
            if len(cell_value) > max_length:
                max_length = min(len(cell_value), 50)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    # BytesIO-д хадгалах
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_sample_export(samples: List, include_results: bool = False) -> BytesIO:
    """
    Дээжний өгөгдлийг Excel экспорт

    Args:
        samples: Sample объектуудын жагсаалт
        include_results: Шинжилгээний үр дүнг оруулах эсэх
    """
    columns = [
        {"key": "id", "label": "ID"},
        {"key": "sample_code", "label": "Дээжний код"},
        {"key": "client_name", "label": "Нэгж"},
        {"key": "sample_type", "label": "Төрөл"},
        {"key": "sample_date", "label": "Дээжний огноо"},
        {"key": "received_date", "label": "Хүлээн авсан"},
        {"key": "status", "label": "Статус"},
        {"key": "delivered_by", "label": "Хүлээлгэж өгсөн"},
    ]

    data = []
    for s in samples:
        row = {
            "id": s.id,
            "sample_code": s.sample_code,
            "client_name": s.client_name,
            "sample_type": s.sample_type,
            "sample_date": s.sample_date.strftime('%Y-%m-%d') if s.sample_date else '',
            "received_date": s.received_date.strftime('%Y-%m-%d %H:%M') if s.received_date else '',
            "status": s.status,
            "delivered_by": s.delivered_by or '',
        }
        data.append(row)

    return export_to_excel(data, columns, "samples_export.xlsx", "Samples")


def create_analysis_export(results: List) -> BytesIO:
    """
    Шинжилгээний үр дүнг Excel экспорт
    """
    columns = [
        {"key": "id", "label": "ID"},
        {"key": "sample_code", "label": "Дээжний код"},
        {"key": "analysis_code", "label": "Шинжилгээ"},
        {"key": "final_result", "label": "Үр дүн"},
        {"key": "status", "label": "Статус"},
        {"key": "operator", "label": "Химич"},
        {"key": "created_at", "label": "Огноо"},
    ]

    data = []
    for r in results:
        row = {
            "id": r.id,
            "sample_code": r.sample.sample_code if r.sample else '',
            "analysis_code": r.analysis_code,
            "final_result": r.final_result,
            "status": r.status,
            "operator": r.user.username if r.user else '',
            "created_at": r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
        }
        data.append(row)

    return export_to_excel(data, columns, "analysis_export.xlsx", "Analysis Results")


def create_audit_export(logs: List) -> BytesIO:
    """
    Аудит логийг Excel экспорт
    """
    columns = [
        {"key": "id", "label": "ID"},
        {"key": "timestamp", "label": "Огноо"},
        {"key": "user", "label": "Хэрэглэгч"},
        {"key": "action", "label": "Үйлдэл"},
        {"key": "resource_type", "label": "Төрөл"},
        {"key": "resource_id", "label": "ID"},
        {"key": "ip_address", "label": "IP"},
        {"key": "details", "label": "Дэлгэрэнгүй"},
    ]

    data = []
    for log in logs:
        row = {
            "id": log.id,
            "timestamp": log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
            "user": log.user.username if log.user else '',
            "action": log.action,
            "resource_type": log.resource_type or '',
            "resource_id": log.resource_id or '',
            "ip_address": log.ip_address or '',
            "details": str(log.details)[:500] if log.details else '',
        }
        data.append(row)

    return export_to_excel(data, columns, "audit_export.xlsx", "Audit Log")


def send_excel_response(excel_data: BytesIO, filename: str):
    """
    Excel файлыг HTTP response болгож буцаах
    """
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
