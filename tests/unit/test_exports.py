# -*- coding: utf-8 -*-
"""
Tests for app/utils/exports.py
Excel export utility tests
"""
import pytest
from io import BytesIO
from datetime import datetime
from unittest.mock import patch, MagicMock


class TestExportToExcel:
    """export_to_excel function tests"""

    def test_import(self):
        """Import function"""
        from app.utils.exports import export_to_excel
        assert callable(export_to_excel)

    def test_returns_bytesio(self):
        """Returns BytesIO"""
        from app.utils.exports import export_to_excel

        data = [{"col1": "value1"}]
        columns = [{"key": "col1", "label": "Column 1"}]

        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_empty_data(self):
        """Handles empty data"""
        from app.utils.exports import export_to_excel

        data = []
        columns = [{"key": "col1", "label": "Column 1"}]

        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_single_row(self):
        """Single row export"""
        from app.utils.exports import export_to_excel

        data = [{"name": "Test", "value": 123}]
        columns = [
            {"key": "name", "label": "Name"},
            {"key": "value", "label": "Value"}
        ]

        result = export_to_excel(data, columns)
        assert result.getvalue()  # Has content

    def test_multiple_rows(self):
        """Multiple rows export"""
        from app.utils.exports import export_to_excel

        data = [
            {"name": "Row1", "value": 1},
            {"name": "Row2", "value": 2},
            {"name": "Row3", "value": 3}
        ]
        columns = [
            {"key": "name", "label": "Name"},
            {"key": "value", "label": "Value"}
        ]

        result = export_to_excel(data, columns)
        assert result.getvalue()

    def test_missing_key_in_data(self):
        """Handles missing key in data"""
        from app.utils.exports import export_to_excel

        data = [{"name": "Test"}]  # No 'value' key
        columns = [
            {"key": "name", "label": "Name"},
            {"key": "value", "label": "Value"}
        ]

        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_custom_sheet_name(self):
        """Custom sheet name"""
        from app.utils.exports import export_to_excel
        from openpyxl import load_workbook

        data = [{"col1": "value1"}]
        columns = [{"key": "col1", "label": "Column 1"}]

        result = export_to_excel(data, columns, sheet_name="CustomSheet")

        wb = load_workbook(result)
        assert "CustomSheet" in wb.sheetnames

    def test_header_styling(self):
        """Headers have styling"""
        from app.utils.exports import export_to_excel
        from openpyxl import load_workbook

        data = [{"col1": "value1"}]
        columns = [{"key": "col1", "label": "Column 1"}]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)

        assert header_cell.font.bold is True

    def test_long_cell_value(self):
        """Handles long cell values"""
        from app.utils.exports import export_to_excel

        long_text = "A" * 1000
        data = [{"text": long_text}]
        columns = [{"key": "text", "label": "Text"}]

        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_numeric_values(self):
        """Handles numeric values"""
        from app.utils.exports import export_to_excel
        from openpyxl import load_workbook

        data = [{"num": 123.456}]
        columns = [{"key": "num", "label": "Number"}]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 123.456

    def test_none_values(self):
        """Handles None values"""
        from app.utils.exports import export_to_excel

        data = [{"col1": None}]
        columns = [{"key": "col1", "label": "Column 1"}]

        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)


class TestCreateSampleExport:
    """create_sample_export function tests"""

    def test_import(self):
        """Import function"""
        from app.utils.exports import create_sample_export
        assert callable(create_sample_export)

    def test_returns_bytesio(self):
        """Returns BytesIO"""
        from app.utils.exports import create_sample_export

        mock_sample = MagicMock()
        mock_sample.id = 1
        mock_sample.sample_code = "S001"
        mock_sample.client_name = "Client"
        mock_sample.sample_type = "Coal"
        mock_sample.sample_date = datetime(2025, 1, 1)
        mock_sample.received_date = datetime(2025, 1, 2)
        mock_sample.status = "registered"
        mock_sample.delivered_by = "John"

        result = create_sample_export([mock_sample])
        assert isinstance(result, BytesIO)

    def test_empty_samples(self):
        """Handles empty samples list"""
        from app.utils.exports import create_sample_export

        result = create_sample_export([])
        assert isinstance(result, BytesIO)

    def test_sample_without_dates(self):
        """Handles sample without dates"""
        from app.utils.exports import create_sample_export

        mock_sample = MagicMock()
        mock_sample.id = 1
        mock_sample.sample_code = "S001"
        mock_sample.client_name = "Client"
        mock_sample.sample_type = "Coal"
        mock_sample.sample_date = None
        mock_sample.received_date = None
        mock_sample.status = "registered"
        mock_sample.delivered_by = None

        result = create_sample_export([mock_sample])
        assert isinstance(result, BytesIO)

    def test_multiple_samples(self):
        """Multiple samples export"""
        from app.utils.exports import create_sample_export

        samples = []
        for i in range(5):
            mock_sample = MagicMock()
            mock_sample.id = i
            mock_sample.sample_code = f"S00{i}"
            mock_sample.client_name = "Client"
            mock_sample.sample_type = "Coal"
            mock_sample.sample_date = datetime(2025, 1, 1)
            mock_sample.received_date = datetime(2025, 1, 2)
            mock_sample.status = "registered"
            mock_sample.delivered_by = "John"
            samples.append(mock_sample)

        result = create_sample_export(samples)
        assert result.getvalue()

    def test_include_results_parameter(self, app):
        """Accepts include_results parameter"""
        from app.utils.exports import create_sample_export

        with app.app_context():
            mock_sample = MagicMock()
            mock_sample.id = 1
            mock_sample.sample_code = "S001"
            mock_sample.client_name = "Client"
            mock_sample.sample_type = "Coal"
            mock_sample.sample_date = None
            mock_sample.received_date = None
            mock_sample.status = "registered"
            mock_sample.delivered_by = None

            # Should not raise error
            result = create_sample_export([mock_sample], _include_results=True)
            assert isinstance(result, BytesIO)


class TestCreateAnalysisExport:
    """create_analysis_export function tests"""

    def test_import(self):
        """Import function"""
        from app.utils.exports import create_analysis_export
        assert callable(create_analysis_export)

    def test_returns_bytesio(self):
        """Returns BytesIO"""
        from app.utils.exports import create_analysis_export

        mock_sample = MagicMock()
        mock_sample.sample_code = "S001"

        mock_user = MagicMock()
        mock_user.username = "operator"

        mock_result = MagicMock()
        mock_result.id = 1
        mock_result.sample = mock_sample
        mock_result.analysis_code = "MAD"
        mock_result.final_result = 5.5
        mock_result.status = "approved"
        mock_result.user = mock_user
        mock_result.created_at = datetime(2025, 1, 1)

        result = create_analysis_export([mock_result])
        assert isinstance(result, BytesIO)

    def test_empty_results(self):
        """Handles empty results list"""
        from app.utils.exports import create_analysis_export

        result = create_analysis_export([])
        assert isinstance(result, BytesIO)

    def test_result_without_sample(self):
        """Handles result without sample"""
        from app.utils.exports import create_analysis_export

        mock_result = MagicMock()
        mock_result.id = 1
        mock_result.sample = None
        mock_result.analysis_code = "MAD"
        mock_result.final_result = 5.5
        mock_result.status = "approved"
        mock_result.user = None
        mock_result.created_at = None

        result = create_analysis_export([mock_result])
        assert isinstance(result, BytesIO)

    def test_multiple_results(self):
        """Multiple results export"""
        from app.utils.exports import create_analysis_export

        results = []
        for i in range(5):
            mock_sample = MagicMock()
            mock_sample.sample_code = f"S00{i}"

            mock_result = MagicMock()
            mock_result.id = i
            mock_result.sample = mock_sample
            mock_result.analysis_code = "MAD"
            mock_result.final_result = 5.5 + i
            mock_result.status = "approved"
            mock_result.user = None
            mock_result.created_at = datetime(2025, 1, 1)
            results.append(mock_result)

        result = create_analysis_export(results)
        assert result.getvalue()


class TestCreateAuditExport:
    """create_audit_export function tests"""

    def test_import(self):
        """Import function"""
        from app.utils.exports import create_audit_export
        assert callable(create_audit_export)

    def test_returns_bytesio(self):
        """Returns BytesIO"""
        from app.utils.exports import create_audit_export

        mock_user = MagicMock()
        mock_user.username = "admin"

        mock_log = MagicMock()
        mock_log.id = 1
        mock_log.timestamp = datetime(2025, 1, 1, 12, 0, 0)
        mock_log.user = mock_user
        mock_log.action = "CREATE"
        mock_log.resource_type = "Sample"
        mock_log.resource_id = 1
        mock_log.ip_address = "127.0.0.1"
        mock_log.details = {"key": "value"}

        result = create_audit_export([mock_log])
        assert isinstance(result, BytesIO)

    def test_empty_logs(self):
        """Handles empty logs list"""
        from app.utils.exports import create_audit_export

        result = create_audit_export([])
        assert isinstance(result, BytesIO)

    def test_log_without_user(self):
        """Handles log without user"""
        from app.utils.exports import create_audit_export

        mock_log = MagicMock()
        mock_log.id = 1
        mock_log.timestamp = None
        mock_log.user = None
        mock_log.action = "CREATE"
        mock_log.resource_type = None
        mock_log.resource_id = None
        mock_log.ip_address = None
        mock_log.details = None

        result = create_audit_export([mock_log])
        assert isinstance(result, BytesIO)

    def test_long_details(self):
        """Handles long details (truncated to 500)"""
        from app.utils.exports import create_audit_export

        mock_log = MagicMock()
        mock_log.id = 1
        mock_log.timestamp = datetime(2025, 1, 1)
        mock_log.user = None
        mock_log.action = "CREATE"
        mock_log.resource_type = "Sample"
        mock_log.resource_id = 1
        mock_log.ip_address = "127.0.0.1"
        mock_log.details = "A" * 1000  # Long details

        result = create_audit_export([mock_log])
        assert isinstance(result, BytesIO)

    def test_multiple_logs(self):
        """Multiple logs export"""
        from app.utils.exports import create_audit_export

        logs = []
        for i in range(10):
            mock_log = MagicMock()
            mock_log.id = i
            mock_log.timestamp = datetime(2025, 1, 1)
            mock_log.user = None
            mock_log.action = "CREATE"
            mock_log.resource_type = "Sample"
            mock_log.resource_id = i
            mock_log.ip_address = "127.0.0.1"
            mock_log.details = None
            logs.append(mock_log)

        result = create_audit_export(logs)
        assert result.getvalue()


class TestSendExcelResponse:
    """send_excel_response function tests"""

    def test_import(self):
        """Import function"""
        from app.utils.exports import send_excel_response
        assert callable(send_excel_response)

    def test_calls_send_file(self):
        """Calls send_file"""
        from app.utils.exports import send_excel_response

        with patch('app.utils.exports.send_file') as mock_send:
            mock_send.return_value = "response"

            excel_data = BytesIO(b"test data")
            result = send_excel_response(excel_data, "test.xlsx")

            mock_send.assert_called_once()
            assert result == "response"

    def test_correct_mimetype(self):
        """Uses correct mimetype"""
        from app.utils.exports import send_excel_response

        with patch('app.utils.exports.send_file') as mock_send:
            mock_send.return_value = "response"

            excel_data = BytesIO(b"test data")
            send_excel_response(excel_data, "test.xlsx")

            args, kwargs = mock_send.call_args
            assert kwargs['mimetype'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def test_as_attachment(self):
        """Sets as_attachment to True"""
        from app.utils.exports import send_excel_response

        with patch('app.utils.exports.send_file') as mock_send:
            mock_send.return_value = "response"

            excel_data = BytesIO(b"test data")
            send_excel_response(excel_data, "test.xlsx")

            args, kwargs = mock_send.call_args
            assert kwargs['as_attachment'] is True

    def test_download_name(self):
        """Sets download_name"""
        from app.utils.exports import send_excel_response

        with patch('app.utils.exports.send_file') as mock_send:
            mock_send.return_value = "response"

            excel_data = BytesIO(b"test data")
            send_excel_response(excel_data, "my_export.xlsx")

            args, kwargs = mock_send.call_args
            assert kwargs['download_name'] == "my_export.xlsx"
