# -*- coding: utf-8 -*-
"""
exports.py модулийн 100% coverage тестүүд

Excel экспорт функцүүдийн бүх branch-уудыг тест хийнэ.
"""
import pytest
from unittest.mock import patch, MagicMock
from io import BytesIO
from datetime import datetime, date


class TestExportsImport:
    """Модуль импорт тест"""

    def test_import_module(self):
        from app.utils import exports
        assert exports is not None

    def test_import_functions(self):
        from app.utils.exports import (
            export_to_excel, create_sample_export,
            create_analysis_export, create_audit_export,
            send_excel_response
        )
        assert export_to_excel is not None
        assert create_sample_export is not None
        assert create_analysis_export is not None
        assert create_audit_export is not None
        assert send_excel_response is not None


class TestExportToExcel:
    """export_to_excel функцийн тест"""

    def test_empty_data(self):
        from app.utils.exports import export_to_excel
        data = []
        columns = [{"key": "id", "label": "ID"}]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_single_row(self):
        from app.utils.exports import export_to_excel
        data = [{"id": 1, "name": "Test"}]
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Name"}
        ]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)
        # Check file is not empty
        result.seek(0, 2)  # Seek to end
        assert result.tell() > 0

    def test_multiple_rows(self):
        from app.utils.exports import export_to_excel
        data = [
            {"id": 1, "name": "Test 1"},
            {"id": 2, "name": "Test 2"},
            {"id": 3, "name": "Test 3"}
        ]
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Name"}
        ]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_missing_key(self):
        """Key байхгүй бол хоосон утга"""
        from app.utils.exports import export_to_excel
        data = [{"id": 1}]  # No "name" key
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Name"}
        ]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_custom_sheet_name(self):
        from app.utils.exports import export_to_excel
        data = [{"id": 1}]
        columns = [{"key": "id", "label": "ID"}]
        result = export_to_excel(data, columns, sheet_name="CustomSheet")
        assert isinstance(result, BytesIO)

    def test_custom_filename(self):
        from app.utils.exports import export_to_excel
        data = [{"id": 1}]
        columns = [{"key": "id", "label": "ID"}]
        result = export_to_excel(data, columns, filename="custom.xlsx")
        assert isinstance(result, BytesIO)

    def test_long_cell_value(self):
        """Урт утга (max 50 chars in column width)"""
        from app.utils.exports import export_to_excel
        data = [{"id": 1, "name": "A" * 100}]  # 100 char string
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Name"}
        ]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_numeric_values(self):
        from app.utils.exports import export_to_excel
        data = [{"value": 123.456}]
        columns = [{"key": "value", "label": "Value"}]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)

    def test_none_value(self):
        from app.utils.exports import export_to_excel
        data = [{"id": None}]
        columns = [{"key": "id", "label": "ID"}]
        result = export_to_excel(data, columns)
        assert isinstance(result, BytesIO)


class TestCreateSampleExport:
    """create_sample_export функцийн тест"""

    def test_empty_samples(self):
        from app.utils.exports import create_sample_export
        result = create_sample_export([])
        assert isinstance(result, BytesIO)

    def test_single_sample(self):
        from app.utils.exports import create_sample_export
        sample = MagicMock()
        sample.id = 1
        sample.sample_code = "SAMPLE-001"
        sample.client_name = "Client A"
        sample.sample_type = "Coal"
        sample.sample_date = date(2026, 5, 15)
        sample.received_date = datetime(2026, 5, 15, 10, 30)
        sample.status = "approved"
        sample.delivered_by = "John Doe"

        result = create_sample_export([sample])
        assert isinstance(result, BytesIO)

    def test_sample_with_none_dates(self):
        from app.utils.exports import create_sample_export
        sample = MagicMock()
        sample.id = 1
        sample.sample_code = "SAMPLE-001"
        sample.client_name = "Client A"
        sample.sample_type = "Coal"
        sample.sample_date = None
        sample.received_date = None
        sample.status = "pending"
        sample.delivered_by = None

        result = create_sample_export([sample])
        assert isinstance(result, BytesIO)

    def test_multiple_samples(self):
        from app.utils.exports import create_sample_export
        samples = []
        for i in range(5):
            sample = MagicMock()
            sample.id = i + 1
            sample.sample_code = f"SAMPLE-{i+1:03d}"
            sample.client_name = f"Client {i+1}"
            sample.sample_type = "Coal"
            sample.sample_date = date(2026, 5, i+1)
            sample.received_date = datetime(2026, 5, i+1, 10, 0)
            sample.status = "approved"
            sample.delivered_by = f"Person {i+1}"
            samples.append(sample)

        result = create_sample_export(samples)
        assert isinstance(result, BytesIO)

    def test_include_results_false(self):
        from app.utils.exports import create_sample_export
        sample = MagicMock()
        sample.id = 1
        sample.sample_code = "SAMPLE-001"
        sample.client_name = "Client A"
        sample.sample_type = "Coal"
        sample.sample_date = None
        sample.received_date = None
        sample.status = "pending"
        sample.delivered_by = None

        result = create_sample_export([sample], _include_results=False)
        assert isinstance(result, BytesIO)


class TestCreateAnalysisExport:
    """create_analysis_export функцийн тест"""

    def test_empty_results(self):
        from app.utils.exports import create_analysis_export
        result = create_analysis_export([])
        assert isinstance(result, BytesIO)

    def test_single_result(self):
        from app.utils.exports import create_analysis_export
        ar = MagicMock()
        ar.id = 1
        ar.sample = MagicMock()
        ar.sample.sample_code = "SAMPLE-001"
        ar.analysis_code = "MT"
        ar.final_result = 5.25
        ar.status = "approved"
        ar.user = MagicMock()
        ar.user.username = "chemist1"
        ar.created_at = datetime(2026, 5, 15, 14, 30)

        result = create_analysis_export([ar])
        assert isinstance(result, BytesIO)

    def test_result_with_none_sample(self):
        from app.utils.exports import create_analysis_export
        ar = MagicMock()
        ar.id = 1
        ar.sample = None
        ar.analysis_code = "MT"
        ar.final_result = 5.25
        ar.status = "approved"
        ar.user = None
        ar.created_at = None

        result = create_analysis_export([ar])
        assert isinstance(result, BytesIO)

    def test_multiple_results(self):
        from app.utils.exports import create_analysis_export
        results = []
        for i in range(5):
            ar = MagicMock()
            ar.id = i + 1
            ar.sample = MagicMock()
            ar.sample.sample_code = f"SAMPLE-{i+1:03d}"
            ar.analysis_code = ["MT", "Aad", "Vad", "CV", "TS"][i]
            ar.final_result = 10.0 + i
            ar.status = "approved"
            ar.user = MagicMock()
            ar.user.username = f"chemist{i+1}"
            ar.created_at = datetime(2026, 5, 15, 10 + i, 0)
            results.append(ar)

        result = create_analysis_export(results)
        assert isinstance(result, BytesIO)


class TestCreateAuditExport:
    """create_audit_export функцийн тест"""

    def test_empty_logs(self):
        from app.utils.exports import create_audit_export
        result = create_audit_export([])
        assert isinstance(result, BytesIO)

    def test_single_log(self):
        from app.utils.exports import create_audit_export
        log = MagicMock()
        log.id = 1
        log.timestamp = datetime(2026, 5, 15, 10, 30, 45)
        log.user = MagicMock()
        log.user.username = "admin"
        log.action = "create"
        log.resource_type = "sample"
        log.resource_id = 123
        log.ip_address = "192.168.1.1"
        log.details = {"sample_code": "SAMPLE-001"}

        result = create_audit_export([log])
        assert isinstance(result, BytesIO)

    def test_log_with_none_values(self):
        from app.utils.exports import create_audit_export
        log = MagicMock()
        log.id = 1
        log.timestamp = None
        log.user = None
        log.action = "login"
        log.resource_type = None
        log.resource_id = None
        log.ip_address = None
        log.details = None

        result = create_audit_export([log])
        assert isinstance(result, BytesIO)

    def test_log_with_long_details(self):
        from app.utils.exports import create_audit_export
        log = MagicMock()
        log.id = 1
        log.timestamp = datetime(2026, 5, 15, 10, 0)
        log.user = MagicMock()
        log.user.username = "admin"
        log.action = "update"
        log.resource_type = "sample"
        log.resource_id = 1
        log.ip_address = "192.168.1.1"
        log.details = {"data": "A" * 1000}  # Long details

        result = create_audit_export([log])
        assert isinstance(result, BytesIO)

    def test_multiple_logs(self):
        from app.utils.exports import create_audit_export
        logs = []
        for i in range(5):
            log = MagicMock()
            log.id = i + 1
            log.timestamp = datetime(2026, 5, 15, 10 + i, 0, 0)
            log.user = MagicMock()
            log.user.username = "admin"
            log.action = ["create", "read", "update", "delete", "login"][i]
            log.resource_type = "sample"
            log.resource_id = i + 1
            log.ip_address = f"192.168.1.{i+1}"
            log.details = {"index": i}
            logs.append(log)

        result = create_audit_export(logs)
        assert isinstance(result, BytesIO)


class TestSendExcelResponse:
    """send_excel_response функцийн тест"""

    @patch('app.utils.exports.send_file')
    def test_send_excel_response(self, mock_send_file):
        from app.utils.exports import send_excel_response
        mock_send_file.return_value = "response"

        excel_data = BytesIO(b"test data")
        result = send_excel_response(excel_data, "test.xlsx")

        mock_send_file.assert_called_once()
        call_args = mock_send_file.call_args
        assert call_args[1]['mimetype'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert call_args[1]['as_attachment'] is True
        assert call_args[1]['download_name'] == 'test.xlsx'

    @patch('app.utils.exports.send_file')
    def test_send_excel_response_custom_filename(self, mock_send_file):
        from app.utils.exports import send_excel_response
        mock_send_file.return_value = "response"

        excel_data = BytesIO(b"test data")
        result = send_excel_response(excel_data, "custom_export_2026.xlsx")

        call_args = mock_send_file.call_args
        assert call_args[1]['download_name'] == 'custom_export_2026.xlsx'


class TestExcelContentValidation:
    """Excel файлын агуулгыг шалгах"""

    def test_excel_can_be_opened(self):
        """Үүссэн Excel файлыг openpyxl-ээр нээж болно"""
        from app.utils.exports import export_to_excel
        from openpyxl import load_workbook

        data = [
            {"id": 1, "name": "Test 1"},
            {"id": 2, "name": "Test 2"}
        ]
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Нэр"}  # Mongolian text
        ]

        result = export_to_excel(data, columns, sheet_name="TestSheet")
        result.seek(0)

        wb = load_workbook(result)
        ws = wb.active

        # Check sheet name
        assert ws.title == "TestSheet"

        # Check header row
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Нэр"

        # Check data rows
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "Test 1"
        assert ws.cell(3, 1).value == 2
        assert ws.cell(3, 2).value == "Test 2"

    def test_excel_styling(self):
        """Excel стиль шалгах"""
        from app.utils.exports import export_to_excel
        from openpyxl import load_workbook

        data = [{"id": 1}]
        columns = [{"key": "id", "label": "ID"}]

        result = export_to_excel(data, columns)
        result.seek(0)

        wb = load_workbook(result)
        ws = wb.active

        # Check header has styling
        header_cell = ws.cell(1, 1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None

    def test_excel_with_unicode(self):
        """Unicode текст"""
        from app.utils.exports import export_to_excel
        from openpyxl import load_workbook

        data = [{"name": "Монгол тест 日本語 한국어"}]
        columns = [{"key": "name", "label": "Нэр"}]

        result = export_to_excel(data, columns)
        result.seek(0)

        wb = load_workbook(result)
        ws = wb.active

        assert "Монгол" in str(ws.cell(2, 1).value)
