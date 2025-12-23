# tests/unit/test_exports_comprehensive.py
# -*- coding: utf-8 -*-
"""
exports.py - Excel export функцүүдийн тест
"""
import pytest
from io import BytesIO
from unittest.mock import Mock, MagicMock
from datetime import datetime
from openpyxl import load_workbook


class TestExportToExcel:
    """export_to_excel функцийн тестүүд"""

    def test_export_empty_data(self):
        """Хоосон өгөгдөл экспорт"""
        from app.utils.exports import export_to_excel

        data = []
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Нэр"}
        ]

        result = export_to_excel(data, columns)

        assert isinstance(result, BytesIO)
        # Excel файл уншиж шалгах
        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=1, column=2).value == "Нэр"
        # Өгөгдөл байхгүй
        assert ws.cell(row=2, column=1).value is None

    def test_export_with_data(self):
        """Өгөгдөлтэй экспорт"""
        from app.utils.exports import export_to_excel

        data = [
            {"id": 1, "name": "Тест 1", "value": 100},
            {"id": 2, "name": "Тест 2", "value": 200},
        ]
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Нэр"},
            {"key": "value", "label": "Утга"}
        ]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        # Толгой
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=1, column=2).value == "Нэр"
        assert ws.cell(row=1, column=3).value == "Утга"
        # Өгөгдөл
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "Тест 1"
        assert ws.cell(row=2, column=3).value == 100
        assert ws.cell(row=3, column=1).value == 2

    def test_export_custom_sheet_name(self):
        """Sheet нэр өөрчлөх"""
        from app.utils.exports import export_to_excel

        data = [{"id": 1}]
        columns = [{"key": "id", "label": "ID"}]

        result = export_to_excel(data, columns, sheet_name="CustomSheet")

        wb = load_workbook(result)
        assert wb.active.title == "CustomSheet"

    def test_export_missing_keys(self):
        """Дутуу key байхад хоосон утга"""
        from app.utils.exports import export_to_excel

        data = [
            {"id": 1, "name": "Тест"},
            {"id": 2},  # name байхгүй
        ]
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Нэр"}
        ]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "Тест"
        # openpyxl хоосон string-г None болгодог
        assert ws.cell(row=3, column=2).value in ("", None)

    def test_export_long_text(self):
        """Урт текст багтаах"""
        from app.utils.exports import export_to_excel

        long_text = "A" * 100
        data = [{"text": long_text}]
        columns = [{"key": "text", "label": "Текст"}]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == long_text

    def test_export_special_characters(self):
        """Тусгай тэмдэгтүүд"""
        from app.utils.exports import export_to_excel

        data = [{"text": "Монгол хэл <>&\"'"}]
        columns = [{"key": "text", "label": "Текст"}]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Монгол хэл <>&\"'"

    def test_export_numeric_values(self):
        """Тоон утгууд"""
        from app.utils.exports import export_to_excel

        data = [
            {"int_val": 123, "float_val": 45.67, "negative": -100}
        ]
        columns = [
            {"key": "int_val", "label": "Бүхэл"},
            {"key": "float_val", "label": "Бутархай"},
            {"key": "negative", "label": "Сөрөг"}
        ]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 123
        assert ws.cell(row=2, column=2).value == 45.67
        assert ws.cell(row=2, column=3).value == -100

    def test_export_none_values(self):
        """None утгууд"""
        from app.utils.exports import export_to_excel

        data = [{"val": None}]
        columns = [{"key": "val", "label": "Утга"}]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        # openpyxl хоосон string-г None болгодог
        assert ws.cell(row=2, column=1).value in ("", None)

    def test_export_many_rows(self):
        """Олон мөр"""
        from app.utils.exports import export_to_excel

        data = [{"id": i, "name": f"Row {i}"} for i in range(100)]
        columns = [
            {"key": "id", "label": "ID"},
            {"key": "name", "label": "Нэр"}
        ]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=101, column=1).value == 99
        assert ws.cell(row=101, column=2).value == "Row 99"

    def test_export_many_columns(self):
        """Олон багана"""
        from app.utils.exports import export_to_excel

        data = [{f"col{i}": f"val{i}" for i in range(20)}]
        columns = [{"key": f"col{i}", "label": f"Багана {i}"} for i in range(20)]

        result = export_to_excel(data, columns)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=1, column=20).value == "Багана 19"
        assert ws.cell(row=2, column=20).value == "val19"


class TestCreateSampleExport:
    """create_sample_export функцийн тестүүд"""

    def test_sample_export_basic(self):
        """Энгийн дээж экспорт"""
        from app.utils.exports import create_sample_export

        # Mock sample objects
        sample1 = Mock()
        sample1.id = 1
        sample1.sample_code = "S001"
        sample1.client_name = "Client A"
        sample1.sample_type = "Coal"
        sample1.sample_date = datetime(2024, 1, 15)
        sample1.received_date = datetime(2024, 1, 15, 10, 30)
        sample1.status = "pending"
        sample1.delivered_by = "User1"

        sample2 = Mock()
        sample2.id = 2
        sample2.sample_code = "S002"
        sample2.client_name = "Client B"
        sample2.sample_type = "Water"
        sample2.sample_date = None
        sample2.received_date = None
        sample2.status = "approved"
        sample2.delivered_by = None

        result = create_sample_export([sample1, sample2])

        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "Samples"
        assert ws.cell(row=2, column=2).value == "S001"
        assert ws.cell(row=3, column=2).value == "S002"

    def test_sample_export_empty(self):
        """Хоосон дээж жагсаалт"""
        from app.utils.exports import create_sample_export

        result = create_sample_export([])

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=2, column=1).value is None


class TestCreateAnalysisExport:
    """create_analysis_export функцийн тестүүд"""

    def test_analysis_export_basic(self):
        """Энгийн шинжилгээ экспорт"""
        from app.utils.exports import create_analysis_export

        # Mock result objects
        sample_mock = Mock()
        sample_mock.sample_code = "S001"

        user_mock = Mock()
        user_mock.username = "chemist1"

        result1 = Mock()
        result1.id = 1
        result1.sample = sample_mock
        result1.analysis_code = "Mad"
        result1.final_result = 5.25
        result1.status = "approved"
        result1.user = user_mock
        result1.created_at = datetime(2024, 1, 15, 14, 30)

        result2 = Mock()
        result2.id = 2
        result2.sample = None
        result2.analysis_code = "Aad"
        result2.final_result = 12.5
        result2.status = "pending"
        result2.user = None
        result2.created_at = None

        output = create_analysis_export([result1, result2])

        wb = load_workbook(output)
        ws = wb.active
        assert ws.title == "Analysis Results"
        assert ws.cell(row=2, column=3).value == "Mad"
        assert ws.cell(row=3, column=3).value == "Aad"


class TestCreateAuditExport:
    """create_audit_export функцийн тестүүд"""

    def test_audit_export_basic(self):
        """Энгийн аудит экспорт"""
        from app.utils.exports import create_audit_export

        user_mock = Mock()
        user_mock.username = "admin"

        log1 = Mock()
        log1.id = 1
        log1.timestamp = datetime(2024, 1, 15, 10, 0, 0)
        log1.user = user_mock
        log1.action = "CREATE"
        log1.resource_type = "Sample"
        log1.resource_id = "123"
        log1.ip_address = "192.168.1.1"
        log1.details = {"key": "value"}

        log2 = Mock()
        log2.id = 2
        log2.timestamp = None
        log2.user = None
        log2.action = "DELETE"
        log2.resource_type = None
        log2.resource_id = None
        log2.ip_address = None
        log2.details = None

        output = create_audit_export([log1, log2])

        wb = load_workbook(output)
        ws = wb.active
        assert ws.title == "Audit Log"
        assert ws.cell(row=2, column=4).value == "CREATE"
        assert ws.cell(row=3, column=4).value == "DELETE"

    def test_audit_export_long_details(self):
        """Урт details талбар (500 тэмдэгт хүртэл)"""
        from app.utils.exports import create_audit_export

        log = Mock()
        log.id = 1
        log.timestamp = datetime.now()
        log.user = None
        log.action = "UPDATE"
        log.resource_type = "Sample"
        log.resource_id = "1"
        log.ip_address = "127.0.0.1"
        log.details = {"long_text": "X" * 1000}

        output = create_audit_export([log])

        wb = load_workbook(output)
        ws = wb.active
        details_val = ws.cell(row=2, column=8).value
        assert len(details_val) <= 500
